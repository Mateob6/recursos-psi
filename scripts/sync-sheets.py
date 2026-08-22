#!/usr/bin/env python3
"""
Sync Google Sheets → src/data/resources.json

Extrae los datos del spreadsheet de Nelson y genera el JSON tipado
que consume el sitio. Diseñado para correr como cron.

Uso:
    cd ~/Desktop/Proyectos/recursos-psi
    uv run scripts/sync-sheets.py

Requisitos:
    uv add openpyxl requests
"""

import json
import re
import sys
from datetime import datetime
from pathlib import Path

SPREADSHEET_ID = "1dZ_LdrQDxj0qI-tL8UNQhiyuohfgax-BgQ1Oz9GSYAg"
EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/export?format=xlsx"
OUTPUT = Path(__file__).parent.parent / "src" / "data" / "resources.json"
CACHE = Path(__file__).parent / ".cache_spreadsheet.xlsx"


def download():
    import requests
    resp = requests.get(EXPORT_URL, timeout=30)
    resp.raise_for_status()
    CACHE.write_bytes(resp.content)
    print(f"Downloaded {len(resp.content)} bytes")


def clean(val):
    if val is None:
        return ""
    return str(val).strip()


def extract_contacts(text):
    if not text:
        return {"raw": ""}

    contacts = {"raw": text}

    wa_links = re.findall(r"https?://wa\.link/\w+", text)
    wa_api = re.findall(r"https?://api\.whatsapp\.com/send/?\?phone=(\d+)", text)
    if wa_links or wa_api:
        contacts["whatsapp"] = wa_links + [
            f"https://api.whatsapp.com/send/?phone={p}" for p in wa_api
        ]

    phones = re.findall(r"\(?\d{3}\)?\s*\d{3}\s*\d{4}", text)
    phones += re.findall(r"\(\d{3}\)\s*\d{7,8}", text)
    if phones:
        contacts["phones"] = list(set(phones))

    emails = re.findall(r"mailto:([^\s,]+)", text)
    if not emails:
        emails = re.findall(r"[\w.+-]+@[\w-]+\.[\w.]+", text)
    if emails:
        contacts["emails"] = list(set(emails))

    urls = re.findall(r"https?://[^\s,]+", text)
    urls = [
        u
        for u in urls
        if "wa.link" not in u
        and "api.whatsapp" not in u
        and "instagram.com" not in u
        and "mailto:" not in u
    ]
    if urls:
        contacts["urls"] = urls

    ig = re.findall(r"instagram\.com/[\w._]+", text)
    if ig:
        contacts["instagram"] = [f"https://www.{i}" for i in ig]

    return contacts


def parse():
    import openpyxl

    wb = openpyxl.load_workbook(CACHE)
    resources = []
    rid = 0

    # 1. Recursos Psicosocial
    ws = wb["Recursos Psicosocial"]
    section = "servicios"
    for row in range(3, ws.max_row + 1):
        a = clean(ws.cell(row=row, column=1).value)
        if not a:
            continue
        if "LÍNEAS DE EMERGENCIA" in a.upper():
            section = "lineas"
            continue
        if section == "lineas" and a == "Entidad":
            continue
        rid += 1
        if section == "servicios":
            resources.append({
                "id": f"psi-{rid}",
                "category": "psicosocial",
                "name": a,
                "center": clean(ws.cell(row=row, column=2).value),
                "modality": clean(ws.cell(row=row, column=3).value),
                "department": clean(ws.cell(row=row, column=4).value),
                "city": clean(ws.cell(row=row, column=5).value),
                "contact": extract_contacts(clean(ws.cell(row=row, column=6).value)),
                "cost": clean(ws.cell(row=row, column=7).value),
                "condition": clean(ws.cell(row=row, column=8).value),
                "serviceType": clean(ws.cell(row=row, column=9).value),
                "targetPopulation": clean(ws.cell(row=row, column=10).value),
                "source": clean(ws.cell(row=row, column=11).value),
                "recommendation": clean(ws.cell(row=row, column=12).value),
            })
        else:
            resources.append({
                "id": f"linea-{rid}",
                "category": "lineas_emergencia",
                "name": a,
                "city": clean(ws.cell(row=row, column=4).value),
                "contact": extract_contacts(clean(ws.cell(row=row, column=6).value)),
                "serviceType": clean(ws.cell(row=row, column=7).value),
            })

    # 2. Recursos de Salud
    ws = wb["Recursos de Salud"]
    section = "eps"
    for row in range(3, ws.max_row + 1):
        a = clean(ws.cell(row=row, column=1).value)
        if not a:
            continue
        if "REGIMENES ESPECIALES" in a.upper():
            section = "regimenes"
            continue
        if "REDES A NIVEL" in a.upper():
            section = "redes"
            continue
        rid += 1
        resources.append({
            "id": f"salud-{rid}",
            "category": "salud",
            "subcategory": section,
            "name": a,
            "contact": extract_contacts(
                clean(ws.cell(row=row, column=2).value) + " " +
                clean(ws.cell(row=row, column=3).value) + " " +
                clean(ws.cell(row=row, column=4).value)
            ),
            "phone": clean(ws.cell(row=row, column=2).value),
            "email": clean(ws.cell(row=row, column=3).value),
            "url": clean(ws.cell(row=row, column=4).value),
            "userNote": clean(ws.cell(row=row, column=5).value),
        })

    # 3. Atención Primaria en Salud
    ws = wb["Atención Primaria en Salud"]
    for row in range(3, ws.max_row + 1):
        a = clean(ws.cell(row=row, column=1).value)
        if not a:
            continue
        rid += 1
        resources.append({
            "id": f"aps-{rid}",
            "category": "atencion_primaria",
            "modality": a,
            "name": clean(ws.cell(row=row, column=2).value),
            "address": clean(ws.cell(row=row, column=3).value),
            "contact": extract_contacts(clean(ws.cell(row=row, column=4).value)),
            "serviceType": clean(ws.cell(row=row, column=5).value),
            "city": "Cali",
        })

    # 4. Recursos de Capacitación
    ws = wb["Recursos de Capacitación"]
    for row in range(3, ws.max_row + 1):
        a = clean(ws.cell(row=row, column=1).value)
        if not a:
            continue
        rid += 1
        resources.append({
            "id": f"cap-{rid}",
            "category": "capacitacion",
            "name": a,
            "description": clean(ws.cell(row=row, column=2).value),
            "contact": extract_contacts(clean(ws.cell(row=row, column=3).value)),
        })

    # 5. Páginas Interactivas
    ws = wb["Páginas Interactivas"]
    for row in range(3, ws.max_row + 1):
        a = clean(ws.cell(row=row, column=1).value)
        if not a:
            continue
        rid += 1
        resources.append({
            "id": f"web-{rid}",
            "category": "interactivas",
            "name": a,
            "contact": extract_contacts(clean(ws.cell(row=row, column=2).value)),
            "description": clean(ws.cell(row=row, column=3).value),
        })

    # 6. Albergues Oficiales
    ws = wb["Albergues Oficiales"]
    section = "oficial"
    for row in range(3, ws.max_row + 1):
        a = clean(ws.cell(row=row, column=1).value)
        if not a or a == "LUGAR":
            continue
        if "ALBERGUES DE LA COMUNIDAD" in a.upper():
            section = "comunitario"
            continue
        rid += 1
        resources.append({
            "id": f"alb-{rid}",
            "category": "albergues",
            "subcategory": section,
            "name": a,
            "city": clean(ws.cell(row=row, column=2).value) or "Cali",
            "address": clean(ws.cell(row=row, column=3).value),
            "contact": extract_contacts(clean(ws.cell(row=row, column=4).value)),
        })

    # 7. Puntos de Acopio
    ws = wb["Puntos de Acopio"]
    for row in range(3, ws.max_row + 1):
        a = clean(ws.cell(row=row, column=1).value)
        if not a:
            continue
        rid += 1
        resources.append({
            "id": f"acopio-{rid}",
            "category": "acopio",
            "name": a,
            "status": clean(ws.cell(row=row, column=2).value),
            "address": clean(ws.cell(row=row, column=3).value),
            "city": "Cali",
        })

    # 8. Servicios Funerarios
    ws = wb["Servicios Funerarios"]
    for row in range(3, ws.max_row + 1):
        a = clean(ws.cell(row=row, column=1).value)
        if not a:
            continue
        rid += 1
        resources.append({
            "id": f"fun-{rid}",
            "category": "funerarios",
            "name": a,
            "status": clean(ws.cell(row=row, column=2).value),
            "address": clean(ws.cell(row=row, column=3).value),
            "requirements": clean(ws.cell(row=row, column=4).value),
            "contact": extract_contacts(clean(ws.cell(row=row, column=5).value)),
        })

    return resources


CATEGORY_TO_SECTION = {
    "psicosocial": "apoyo-emocional",
    "lineas_emergencia": "apoyo-emocional",
    "salud": "salud",
    "atencion_primaria": "salud",
    "albergues": "refugio",
    "acopio": "donaciones",
    "capacitacion": "guias",
    "interactivas": "guias",
    "funerarios": "funerarios",
}

SALUD_SUBCATEGORY_TO_TAG = {
    "eps": "eps",
    "regimenes": "regimen_especial",
    "redes": "informacion",
}

POBLACION_KEYWORDS = [
    ("lgbtiq", "lgbtiq"),
    ("lgtbi", "lgbtiq"),
    ("diversidad sexual", "lgbtiq"),
    ("mujer", "mujeres"),
    ("género", "mujeres"),
    ("niñ", "ninez"),
    ("adolescente", "ninez"),
    ("persona mayor", "persona_mayor"),
    ("adulto mayor", "persona_mayor"),
    ("discapacidad", "discapacidad"),
    ("profesional", "profesionales"),
    ("víctima", "victimas_conflicto"),
    ("conflicto armado", "victimas_conflicto"),
]


def derive_tags(resource):
    cat = resource.get("category", "")
    section = CATEGORY_TO_SECTION.get(cat)
    if not section:
        return {}

    if section == "apoyo-emocional":
        return _tags_apoyo(resource)
    if section == "salud":
        return _tags_salud(resource)
    if section == "refugio":
        return {"tipo": resource.get("subcategory", "oficial")}
    if section == "donaciones":
        status = resource.get("status", "")
        return {"estado": status} if status else {}
    if section == "guias":
        return _tags_guias(resource)
    return {}


def _tags_apoyo(resource):
    cat = resource["category"]
    contact = resource.get("contact", {})
    condition = (resource.get("condition") or "").lower()
    city = (resource.get("city") or "").lower()
    dept = (resource.get("department") or "").lower()
    raw = contact.get("raw", "").lower()
    service = (resource.get("serviceType") or "").lower()
    target = (resource.get("targetPopulation") or "").lower()
    modality = (resource.get("modality") or "").lower()

    urgencia = "ahora" if cat == "lineas_emergencia" else (
        "agendar" if "agendamiento" in condition or "agendar" in condition else "ahora"
    )

    canales = []
    if contact.get("whatsapp"):
        canales.append("whatsapp")
    if contact.get("phones") or "teléfono" in raw or "telefono" in raw or "telefónic" in modality or "marcar" in raw or "indicativo" in raw:
        canales.append("telefono")
    if contact.get("emails"):
        canales.append("correo")
    if "forms.gle" in raw or "formulario" in raw or "docs.google.com/forms" in raw:
        canales.append("formulario")
    if "presencial" in modality or resource.get("address"):
        canales.append("presencial")
    if not canales and contact.get("urls"):
        canales.append("formulario")

    loc = city + " " + dept
    if "nacional" in loc:
        cobertura = "nacional"
    elif "cali" in loc:
        cobertura = "cali"
    else:
        cobertura = "otra"

    text = target + " " + service
    poblacion = []
    seen = set()
    for keyword, tag in POBLACION_KEYWORDS:
        if keyword in text and tag not in seen:
            poblacion.append(tag)
            seen.add(tag)
    if not poblacion:
        poblacion = ["todos"]

    tags = {"urgencia": urgencia, "cobertura": cobertura, "poblacion": poblacion}
    if canales:
        tags["canales"] = canales
    return tags


def _tags_salud(resource):
    cat = resource["category"]
    if cat == "atencion_primaria":
        return {"tipo": "punto_atencion"}
    sub = resource.get("subcategory", "eps")
    return {"tipo": SALUD_SUBCATEGORY_TO_TAG.get(sub, "eps")}


def _tags_guias(resource):
    desc = (resource.get("description") or "").lower()
    name = (resource.get("name") or "").lower()
    contact = resource.get("contact", {})
    urls = contact.get("urls", [])
    cat = resource["category"]
    text = desc + " " + name

    if "profesional" in text or "decálogo" in text:
        audiencia = "profesionales"
    elif cat == "interactivas":
        audiencia = "comunidad"
    else:
        audiencia = "personas_afectadas"

    if any(u.endswith(".pdf") for u in urls) or "pdf" in text:
        formato = "pdf"
    elif any("youtube" in u or "youtu.be" in u for u in urls) or "video" in text or "audiovisual" in text:
        formato = "video"
    else:
        formato = "web"

    return {"audiencia": audiencia, "formato": formato}


def enrich(resources):
    for r in resources:
        r["section"] = CATEGORY_TO_SECTION.get(r.get("category", ""))
        tags = derive_tags(r)
        if tags:
            r["tags"] = tags


def main():
    print("Syncing Google Sheets → resources.json")
    download()
    resources = parse()
    enrich(resources)

    output = {
        "resources": resources,
        "meta": {
            "source": "Google Sheets - Recursos para el apoyo biopsicosocial Univalle Contigo",
            "spreadsheet_id": SPREADSHEET_ID,
            "extracted": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "total": len(resources),
        },
    }

    OUTPUT.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")

    from collections import Counter
    cats = Counter(r["category"] for r in resources)
    print(f"\nTotal: {len(resources)} recursos")
    for cat, count in cats.most_common():
        print(f"  {cat}: {count}")
    print(f"\nSaved to {OUTPUT}")


if __name__ == "__main__":
    main()
